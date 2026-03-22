#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
from datetime import datetime, date
import re
import sys

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data" / "blog_posts.xlsx"


def slugify(text: str) -> str:
    text = (text or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def yaml_quote(value: str) -> str:
    value = (value or "").replace("\\", "\\\\").replace('"', '\\"')
    value = value.replace("\n", "\\n")
    return f'"{value}"'


def as_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_bool_text(value: str) -> bool:
    return as_text(value).lower() in {"1", "true", "yes", "y", "on"}


def load_config(wb):
    ws = wb["Config"]
    config = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        key, value = row
        if key:
            config[str(key).strip()] = "" if value is None else str(value).strip()
    return config


def get_header_map(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip()] = col
    return headers


def cell(ws, row_idx: int, header_map: dict[str, int], header_name: str):
    return ws.cell(row=row_idx, column=header_map[header_name])


def build_front_matter(row_data: dict, config: dict[str, str]) -> str:
    title = row_data["title"]
    publish_date = row_data["publish_date"]
    timezone_offset = config.get("timezone_offset", "+0700")
    layout = row_data.get("layout") or config.get("layout_default", "post")
    category = row_data.get("category") or config.get("default_category", "")
    author = row_data.get("author") or config.get("author_default", "")
    image = row_data.get("image", "")
    excerpt = row_data.get("excerpt", "")
    tags = [t.strip() for t in row_data.get("tags", "").split(",") if t.strip()]

    front = [
        "---",
        f"layout: {layout}",
        f"title: {yaml_quote(title)}",
        f"date: {publish_date} 08:00:00 {timezone_offset}",
    ]

    if category:
        front.append(f"categories: [{category}]")
    if tags:
        front.append("tags: [" + ", ".join(tags) + "]")
    if excerpt:
        front.append(f"excerpt: {yaml_quote(excerpt)}")
    if image:
        field_name = config.get("image_field_name", "image") or "image"
        front.append(f"{field_name}: {yaml_quote(image)}")
    if author:
        field_name = config.get("author_field_name", "author") or "author"
        front.append(f"{field_name}: {yaml_quote(author)}")

    front.extend([
        "---",
        "",
        row_data["content"].rstrip() + "\n",
    ])
    return "\n".join(front)


def main():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH)
    config = load_config(wb)

    posts_directory = config.get("posts_directory", "_posts").strip() or "_posts"
    overwrite_existing = to_bool_text(config.get("overwrite_existing", "no"))
    posts_dir = ROOT / posts_directory
    posts_dir.mkdir(parents=True, exist_ok=True)

    ws = wb["Posts"]
    headers = get_header_map(ws)

    required_headers = [
        "Status", "Publish Date", "Title", "Slug", "Category", "Tags",
        "Excerpt", "Featured Image", "Layout", "Author",
        "Content (Markdown)", "Published At", "Post Path", "Result"
    ]
    missing = [h for h in required_headers if h not in headers]
    if missing:
        raise KeyError(f"Missing headers in Posts sheet: {', '.join(missing)}")

    now = datetime.now()
    changed = False
    created_count = 0

    for row_idx in range(2, ws.max_row + 1):
        status_cell = cell(ws, row_idx, headers, "Status")
        result_cell = cell(ws, row_idx, headers, "Result")
        published_at_cell = cell(ws, row_idx, headers, "Published At")
        post_path_cell = cell(ws, row_idx, headers, "Post Path")
        title_cell = cell(ws, row_idx, headers, "Title")
        slug_cell = cell(ws, row_idx, headers, "Slug")
        publish_date_cell = cell(ws, row_idx, headers, "Publish Date")
        content_cell = cell(ws, row_idx, headers, "Content (Markdown)")

        status = as_text(status_cell.value).lower()
        if status != "ready":
            continue

        title = as_text(title_cell.value)
        slug = as_text(slug_cell.value) or slugify(title)
        if not slug:
            status_cell.value = "error"
            result_cell.value = "Slug kosong dan tidak bisa digenerate dari title."
            changed = True
            continue
        if not as_text(slug_cell.value):
            slug_cell.value = slug
            changed = True

        publish_date_value = publish_date_cell.value
        if publish_date_value in (None, ""):
            publish_date_value = now.date()
            publish_date_cell.value = publish_date_value
            changed = True

        if isinstance(publish_date_value, datetime):
            publish_date = publish_date_value.date().isoformat()
        elif isinstance(publish_date_value, date):
            publish_date = publish_date_value.isoformat()
        else:
            publish_date = as_text(publish_date_value)

        content = as_text(content_cell.value)
        if not title or not content or not publish_date:
            status_cell.value = "error"
            result_cell.value = "Field wajib belum lengkap: Title / Publish Date / Content."
            changed = True
            continue

        row_data = {
            "title": title,
            "slug": slug,
            "publish_date": publish_date,
            "category": as_text(cell(ws, row_idx, headers, "Category").value),
            "tags": as_text(cell(ws, row_idx, headers, "Tags").value),
            "excerpt": as_text(cell(ws, row_idx, headers, "Excerpt").value),
            "image": as_text(cell(ws, row_idx, headers, "Featured Image").value),
            "layout": as_text(cell(ws, row_idx, headers, "Layout").value),
            "author": as_text(cell(ws, row_idx, headers, "Author").value),
            "content": content,
        }

        file_name = f"{publish_date}-{slug}.md"
        output_path = posts_dir / file_name

        if output_path.exists() and not overwrite_existing:
            status_cell.value = "published"
            post_path_cell.value = str(Path(posts_directory) / file_name)
            if not published_at_cell.value:
                published_at_cell.value = now
            result_cell.value = "File sudah ada. Row ditandai sebagai published."
            changed = True
            continue

        markdown = build_front_matter(row_data, config)
        output_path.write_text(markdown, encoding="utf-8")

        status_cell.value = "published"
        published_at_cell.value = now
        post_path_cell.value = str(Path(posts_directory) / file_name)
        result_cell.value = "Published successfully."
        changed = True
        created_count += 1

    if changed:
        wb.save(WORKBOOK_PATH)

    print(f"Done. Created {created_count} post(s). Spreadsheet updated: {changed}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}")
        sys.exit(1)
